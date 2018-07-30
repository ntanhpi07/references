from openpyxl import load_workbook
from openpyxl import Workbook
import glob2
path = 'S:\DDA-DB2\Test\Q1*.xlsx'
owb = Workbook()
ws = owb.active
r=1
for filename in glob2.iglob(path):
    d = filename.find('Q1_')
    inst = filename[d+3:d+7]
    print (filename, inst)
    iwb = load_workbook(filename, read_only=True)
    s = iwb.worksheets[0]
    if not s.min_row==s.max_row:
        firstrow=True
        for row in s.iter_rows():
            if firstrow:
                firstrow=False
                if r>1:
                    continue
            c=2
            for cell in row:
                if c==2:
                    _ = ws.cell(column=1, row=r, value=inst)
                _ = ws.cell(column=c, row=r, value=cell.value)
                c=c+1
            r=r+1
        print(s.title, s.max_row, s.max_column,r)
    owb.save(filename = 'S:\DDA-DB2\Test\db2_q1_2017_merge.xlsx')

